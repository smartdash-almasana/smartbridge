from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from excel_reader import HEADER_ALIASES, _normalize_text, _resolve_column_names
from structured_warnings import (
    w_ambiguous_header_candidates,
    w_header_low_confidence,
    w_header_shifted,
    w_interleaved_empty_columns,
    w_missing_expected_columns,
    w_multiple_table_blocks_possible,
)


def _pick_sheet(excel_path: str, requested_sheet: str | None) -> str:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        if not sheet_names:
            raise ValueError("El archivo no contiene hojas.")

        if requested_sheet is not None:
            if requested_sheet not in sheet_names:
                raise ValueError(f"La hoja '{requested_sheet}' no existe en el archivo.")
            return requested_sheet

        visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
        return visible[0] if visible else sheet_names[0]
    finally:
        wb.close()


def _score_header_candidates(raw_df: pd.DataFrame) -> list[dict[str, object]]:
    candidates: list[dict[str, object]] = []
    for idx, row in raw_df.iterrows():
        normalized_row = {_normalize_text(v) for v in row.tolist()}
        normalized_row.discard("")
        if not normalized_row:
            continue

        matched_by_canonical: dict[str, str] = {}
        for canonical, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                if alias in normalized_row:
                    matched_by_canonical[canonical] = alias
                    break

        candidates.append(
            {
                "row_idx_zero_based": int(idx),
                "score": int(len(matched_by_canonical)),
                "matched": matched_by_canonical,
            }
        )
    return candidates


def _row_has_data(row: pd.Series) -> bool:
    for v in row.tolist():
        if pd.isna(v):
            continue
        if str(v).strip() != "":
            return True
    return False


def detect_tabular_header(excel_path: str, sheet_name: str | None = None) -> dict[str, object]:
    selected_sheet = _pick_sheet(excel_path, sheet_name)
    raw_df = pd.read_excel(excel_path, sheet_name=selected_sheet, header=None)

    warnings: list[dict[str, object]] = []
    candidates = _score_header_candidates(raw_df)
    ranked = sorted(candidates, key=lambda c: (-int(c["score"]), int(c["row_idx_zero_based"])))
    best = ranked[0] if ranked else {"row_idx_zero_based": 0, "score": 0, "matched": {}}
    second = ranked[1] if len(ranked) > 1 else None

    best_score = int(best["score"])
    header_row_zero_based = int(best["row_idx_zero_based"]) if best_score >= 2 else 0

    if best_score >= 3 and (second is None or int(second["score"]) <= 1):
        confianza = "alta"
    elif best_score >= 2:
        confianza = "media"
    else:
        confianza = "baja"

    if best_score < 2:
        warnings.append(w_header_low_confidence(best_score=best_score))

    if second is not None and best_score >= 2 and int(second["score"]) >= 2:
        gap = best_score - int(second["score"])
        if gap <= 1:
            warnings.append(
                w_ambiguous_header_candidates(
                    best_row_1based=int(best["row_idx_zero_based"]) + 1,
                    best_score=best_score,
                    second_row_1based=int(second["row_idx_zero_based"]) + 1,
                    second_score=int(second["score"]),
                )
            )

    if header_row_zero_based > 0:
        warnings.append(w_header_shifted(header_row_1based=header_row_zero_based + 1))

    df = pd.read_excel(excel_path, sheet_name=selected_sheet, header=header_row_zero_based)
    source_columns = [str(c) for c in df.columns.tolist()]
    column_map = _resolve_column_names(list(df.columns))

    columnas_detectadas: dict[str, dict[str, str]] = {}
    for canonical, original in column_map.items():
        normalized = _normalize_text(original)
        matched_alias = normalized if normalized in HEADER_ALIASES.get(canonical, []) else ""
        columnas_detectadas[canonical] = {
            "source_column": str(original),
            "matched_alias": matched_alias,
        }

    expected_columns = ["cliente", "monto", "vencimiento"]
    missing_columns = [col for col in expected_columns if col not in column_map]
    if missing_columns:
        warnings.append(w_missing_expected_columns(missing=missing_columns))

    unnamed_columns = [
        c
        for c in source_columns
        if _normalize_text(c) == "" or _normalize_text(c).startswith("unnamed")
    ]
    if unnamed_columns:
        warnings.append(w_interleaved_empty_columns(columns=unnamed_columns))

    after_header = raw_df.iloc[header_row_zero_based + 1 :].reset_index(drop=True)
    if not after_header.empty:
        empty_run = 0
        first_gap_start: int | None = None
        for i in range(len(after_header)):
            has_data = _row_has_data(after_header.iloc[i])
            if not has_data:
                empty_run += 1
                if empty_run == 1:
                    first_gap_start = i
            else:
                if empty_run >= 2 and first_gap_start is not None:
                    remaining = after_header.iloc[i + 1 :]
                    if any(_row_has_data(remaining.iloc[j]) for j in range(len(remaining))):
                        warnings.append(
                            w_multiple_table_blocks_possible(
                                gap_start_row_1based=header_row_zero_based + 2 + first_gap_start,
                                empty_rows_count=empty_run,
                            )
                        )
                        break
                empty_run = 0
                first_gap_start = None

    return {
        "file": str(Path(excel_path)),
        "sheet": selected_sheet,
        "header_row_1based": int(header_row_zero_based + 1),
        "columnas_detectadas": columnas_detectadas,
        "confianza": confianza,
        "warnings_estructurales": warnings,
    }
