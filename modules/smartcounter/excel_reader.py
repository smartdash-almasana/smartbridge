from datetime import datetime
from pathlib import Path
import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


HEADER_ALIASES = {
    "cliente": [
        "cliente",
        "empresa",
        "razon social",
        "razon_social",
        "nombre cliente",
        "nombre_cliente",
    ],
    "monto": [
        "monto",
        "importe",
        "saldo",
        "deuda",
    ],
    "vencimiento": [
        "vencimiento",
        "fecha vencimiento",
        "fecha_vencimiento",
        "fecha vto",
        "fec vto",
        "f vencimiento",
        "vence",
        "vto",
    ],
}


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _find_header_row(raw_df: pd.DataFrame) -> int:
    for idx, row in raw_df.iterrows():
        normalized_row = {_normalize_text(v) for v in row.tolist()}
        score = 0
        for aliases in HEADER_ALIASES.values():
            if any(alias in normalized_row for alias in aliases):
                score += 1
        if score >= 2:
            return int(idx)
    return 0


def _resolve_column_names(columns: list[object]) -> dict[str, str]:
    resolved: dict[str, str] = {}
    for col in columns:
        norm = _normalize_text(col)
        for canonical, aliases in HEADER_ALIASES.items():
            if canonical in resolved:
                continue
            if norm in aliases:
                resolved[canonical] = str(col)
    return resolved


def _parse_amount(value: object) -> tuple[float | None, str | None]:
    if pd.isna(value):
        return None, None
    if isinstance(value, (int, float)):
        return float(value), None

    raw_text = str(value).strip()
    if not raw_text:
        return None, None

    upper = raw_text.upper()
    if "USD" in upper or "US$" in upper:
        return None, "currency_ambiguous"

    negative = False
    if raw_text.startswith("(") and raw_text.endswith(")"):
        negative = True
        raw_text = raw_text[1:-1].strip()
    if "-" in raw_text:
        negative = True

    text = re.sub(r"[^0-9,.-]", "", raw_text).replace("-", "")
    if not text:
        return None, "non_numeric"

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        if text.count(",") == 1:
            left, right = text.split(",")
            if right.isdigit() and len(right) <= 2:
                text = left + "." + right
            else:
                text = left + right
        else:
            text = text.replace(",", "")
    elif "." in text:
        if text.count(".") == 1:
            left, right = text.split(".")
            if right.isdigit() and len(right) == 3:
                text = left + right
        else:
            parts = text.split(".")
            if all(part.isdigit() for part in parts):
                text = "".join(parts)

    try:
        parsed = float(text)
    except ValueError:
        return None, "non_numeric"

    if negative:
        parsed *= -1
    return parsed, None


def _collect_excel_warnings(
    excel_path: str,
    header_row_zero_based: int,
    df_columns: list[object],
    column_map: dict[str, str],
) -> list[dict[str, object]]:
    warnings: list[dict[str, object]] = []

    wb_formula = load_workbook(excel_path, data_only=False)
    wb_values = load_workbook(excel_path, data_only=True)
    ws_formula = wb_formula[wb_formula.sheetnames[0]]
    ws_values = wb_values[wb_values.sheetnames[0]]

    header_excel_row = header_row_zero_based + 1

    relevant = {}
    for canonical in ("cliente", "monto", "vencimiento"):
        col_name = column_map.get(canonical)
        if not col_name:
            continue
        pos = df_columns.index(col_name)
        col_idx = pos + 1
        relevant[canonical] = col_idx

    hidden_columns = []
    for canonical, col_idx in relevant.items():
        col_letter = get_column_letter(col_idx)
        if ws_formula.column_dimensions[col_letter].hidden:
            hidden_columns.append({"column": col_letter, "field": canonical})
    if hidden_columns:
        warnings.append(
            {
                "code": "hidden_relevant_columns",
                "message": "Hay columnas relevantes ocultas en el Excel.",
                "details": hidden_columns,
            }
        )

    hidden_rows = []
    for row_idx in range(header_excel_row + 1, ws_formula.max_row + 1):
        if ws_formula.row_dimensions[row_idx].hidden:
            for col_idx in relevant.values():
                cell_val = ws_formula.cell(row=row_idx, column=col_idx).value
                if cell_val is not None and str(cell_val).strip() != "":
                    hidden_rows.append(row_idx)
                    break
    if hidden_rows:
        warnings.append(
            {
                "code": "hidden_relevant_rows",
                "message": "Hay filas ocultas con datos en columnas relevantes.",
                "details": {"rows": sorted(set(hidden_rows))},
            }
        )

    formula_issues = []
    for canonical in ("monto", "vencimiento"):
        col_idx = relevant.get(canonical)
        if not col_idx:
            continue
        for row_idx in range(header_excel_row + 1, ws_formula.max_row + 1):
            cell_formula = ws_formula.cell(row=row_idx, column=col_idx).value
            if isinstance(cell_formula, str) and cell_formula.startswith("="):
                cell_value = ws_values.cell(row=row_idx, column=col_idx).value
                if cell_value is None or str(cell_value).strip() == "":
                    formula_issues.append(
                        {
                            "cell": f"{get_column_letter(col_idx)}{row_idx}",
                            "field": canonical,
                            "formula": cell_formula,
                        }
                    )
    if formula_issues:
        warnings.append(
            {
                "code": "formula_without_cached_value",
                "message": "Hay formulas en columnas relevantes sin valor calculado utilizable.",
                "details": formula_issues,
            }
        )

    non_relevant_formula_by_col: dict[int, dict[str, object]] = {}
    relevant_indices = set(relevant.values())
    for col_idx, col_name in enumerate(df_columns, start=1):
        if col_idx in relevant_indices:
            continue
        for row_idx in range(header_excel_row + 1, ws_formula.max_row + 1):
            cell_formula = ws_formula.cell(row=row_idx, column=col_idx).value
            if isinstance(cell_formula, str) and cell_formula.startswith("="):
                if col_idx not in non_relevant_formula_by_col:
                    non_relevant_formula_by_col[col_idx] = {
                        "column": str(col_name),
                        "column_letter": get_column_letter(col_idx),
                        "rows": [],
                    }
                non_relevant_formula_by_col[col_idx]["rows"].append(row_idx)

    if non_relevant_formula_by_col:
        details = []
        total = 0
        for _, item in sorted(non_relevant_formula_by_col.items()):
            rows = sorted(set(item["rows"]))
            count = len(rows)
            total += count
            details.append(
                {
                    "column": item["column"],
                    "column_letter": item["column_letter"],
                    "rows": rows,
                    "count": count,
                }
            )
        warnings.append(
            {
                "code": "formula_non_relevant_detected",
                "message": "Se detectaron fórmulas en columnas no relevantes para el scoring MVP; se registran solo como trazabilidad.",
                "details": {"total_count": total, "columns": details},
            }
        )

    if "color_como_dato" in Path(excel_path).name.lower():
        warnings.append(
            {
                "code": "color_semantics_out_of_scope",
                "message": "Este caso depende de semantica por color y queda fuera de alcance del MVP actual.",
                "details": {},
            }
        )

    return warnings


def get_top_debtors(excel_path: str) -> dict:
    raw_df = pd.read_excel(excel_path, header=None)
    header_row = _find_header_row(raw_df)
    df = pd.read_excel(excel_path, header=header_row)
    source_columns = [str(c) for c in df.columns.tolist()]
    column_map = _resolve_column_names(list(df.columns))

    expected_columns = ["cliente", "monto", "vencimiento"]
    missing_columns = [col for col in expected_columns if col not in column_map]
    if missing_columns:
        raise ValueError(f"Faltan columnas esperadas (normalizadas): {missing_columns}")

    df = df.rename(
        columns={
            column_map["cliente"]: "cliente",
            column_map["monto"]: "monto",
            column_map["vencimiento"]: "vencimiento",
        }
    )
    df = df[["cliente", "monto", "vencimiento"]].copy()
    df["source_row"] = df.index + header_row + 2
    amount_issues: list[dict[str, object]] = []
    parsed_amounts: list[float | None] = []
    for _, row in df.iterrows():
        raw_amount = row["monto"]
        parsed_amount, reason = _parse_amount(raw_amount)
        parsed_amounts.append(parsed_amount)
        if reason is not None:
            amount_issues.append(
                {
                    "row": int(row["source_row"]),
                    "raw_value": str(raw_amount),
                    "reason": reason,
                }
            )
    df["monto"] = parsed_amounts
    df["vencimiento"] = pd.to_datetime(df["vencimiento"], errors="coerce", dayfirst=True)
    df["cliente"] = df["cliente"].astype(str).str.strip()
    df = df[(df["cliente"] != "") & (df["cliente"].str.lower() != "nan")]
    df = df.dropna(subset=["monto", "vencimiento"])

    today = pd.Timestamp(datetime.now().date())
    df["dias_vencido"] = (today - df["vencimiento"]).dt.days

    overdue_df = df[df["dias_vencido"] > 0].copy()
    top_df = overdue_df.sort_values("monto", ascending=False).head(3)

    top_deudores = []
    for _, row in top_df.iterrows():
        top_deudores.append(
            {
                "cliente": str(row["cliente"]),
                "monto": float(row["monto"]),
                "dias_vencido": int(row["dias_vencido"]),
            }
        )

    warnings = _collect_excel_warnings(
        excel_path=excel_path,
        header_row_zero_based=header_row,
        df_columns=source_columns,
        column_map=column_map,
    )
    if amount_issues:
        warnings.append(
            {
                "code": "mixed_amount_parsing_warning",
                "message": "Se detectaron montos con formatos mixtos o moneda ambigua; se usaron solo los valores interpretables para el scoring MVP.",
                "details": {
                    "rows": [item["row"] for item in amount_issues],
                    "raw_values": [item["raw_value"] for item in amount_issues],
                    "issues": amount_issues,
                },
            }
        )

    return {"top_deudores": top_deudores, "warnings": warnings}
