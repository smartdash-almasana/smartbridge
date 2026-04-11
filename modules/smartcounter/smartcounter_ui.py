from __future__ import annotations

import argparse
import io
import json
import os
import re
import shutil
import threading
import webbrowser
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import pandas as pd
from openpyxl import load_workbook

from run_local import run_pipeline_curated


APP_TITLE = "SmartCounter"
APP_SUBTITLE = "Argentum Ledger"

# =============================================================================
# Sección 1: Persistencia local (revisiones + reglas)
# =============================================================================
HTML_PARTS: list[str] = []
REVISIONES_STORE = Path(__file__).resolve().parent / "revisiones_local.json"
REVISIONES_LOCK = threading.Lock()
REGLAS_STORE = Path(__file__).resolve().parent / "reglas_cliente_local.json"
REGLAS_LOCK = threading.Lock()
LOCAL_TMP_ROOT = Path(__file__).resolve().parent / ".smartcounter_tmp"


def _json_response(handler: BaseHTTPRequestHandler, payload: dict, status: int = 200) -> None:
    data = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def _html_response(handler: BaseHTTPRequestHandler, content: str) -> None:
    data = content.encode("utf-8")
    handler.send_response(HTTPStatus.OK)
    handler.send_header("Content-Type", "text/html; charset=utf-8")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


# Persistencia local: Revisiones
def _load_revisiones_store() -> list[dict]:
    if not REVISIONES_STORE.exists():
        return []
    try:
        raw = REVISIONES_STORE.read_text(encoding="utf-8")
        data = json.loads(raw)
        if isinstance(data, list):
            return [item for item in data if isinstance(item, dict)]
        return []
    except Exception:
        return []


def _save_revisiones_store(rows: list[dict]) -> None:
    REVISIONES_STORE.parent.mkdir(parents=True, exist_ok=True)
    temp_path = REVISIONES_STORE.with_suffix(".tmp")
    temp_path.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    temp_path.replace(REVISIONES_STORE)


def _persist_revision(summary: dict) -> dict:
    row = {
        "archivo": summary.get("file"),
        "fecha_hora": datetime.now().isoformat(timespec="seconds"),
        "hoja_detectada": summary.get("sheet"),
        "fila_header": summary.get("header_row_1based"),
        "warning_codes": summary.get("warning_codes", []),
        "resumen_corto": summary.get("resumen_corto"),
    }
    if summary.get("warning_details"):
        row["warning_details"] = summary.get("warning_details")

    with REVISIONES_LOCK:
        rows = _load_revisiones_store()
        row["revision_id"] = len(rows) + 1
        rows.append(row)
        _save_revisiones_store(rows)
    return row


def _list_revisiones(limit: int = 100) -> list[dict]:
    with REVISIONES_LOCK:
        rows = _load_revisiones_store()
    if limit > 0:
        return list(reversed(rows[-limit:]))
    return list(reversed(rows))


def _count_revisiones() -> int:
    with REVISIONES_LOCK:
        return len(_load_revisiones_store())


# Persistencia local: Reglas de cliente
def _default_rules_seed() -> list[dict]:
    now = datetime.now().isoformat(timespec="seconds")
    return [
        {
            "id": 1,
            "cliente": "Cliente general",
            "nombre": "Validación IVA-21",
            "categoria": "Impuestos",
            "origen": "Manual",
            "estado": "Activa",
            "ultimo_uso": now,
            "tipo": "Validación",
            "condicion": "si alícuota IVA es 21%",
            "accion": "normalizar al campo iva_debito",
            "ejemplos": ["IVA 21", "21%", "alicuota_21"],
            "impacto_esperado": "Reduce ambigüedad en impuestos.",
        },
        {
            "id": 2,
            "cliente": "Cliente general",
            "nombre": "Cruce Proveedores A-Z",
            "categoria": "Logística",
            "origen": "Sugerida",
            "estado": "Activa",
            "ultimo_uso": now,
            "tipo": "Matching",
            "condicion": "si CUIT coincide con padrón de proveedor",
            "accion": "asignar proveedor homologado",
            "ejemplos": ["30-12345678-9", "Proveedor ABC"],
            "impacto_esperado": "Mejora correspondencia semántica.",
        },
    ]


def _load_reglas_store() -> list[dict]:
    if not REGLAS_STORE.exists():
        rows = _default_rules_seed()
        _save_reglas_store(rows)
        return rows
    try:
        raw = REGLAS_STORE.read_text(encoding="utf-8")
        data = json.loads(raw)
        if isinstance(data, list):
            return [item for item in data if isinstance(item, dict)]
        return []
    except Exception:
        return []


def _save_reglas_store(rows: list[dict]) -> None:
    REGLAS_STORE.parent.mkdir(parents=True, exist_ok=True)
    temp_path = REGLAS_STORE.with_suffix(".tmp")
    temp_path.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    temp_path.replace(REGLAS_STORE)


def _list_reglas(cliente: str | None = None) -> list[dict]:
    with REGLAS_LOCK:
        rows = _load_reglas_store()
    if cliente:
        filtered = [r for r in rows if str(r.get("cliente", "")).strip().lower() == cliente.strip().lower()]
        if filtered:
            return filtered
    return rows


def _apply_regla_action(action: str, rule_id: int, payload: dict | None = None) -> dict:
    payload = payload or {}
    with REGLAS_LOCK:
        rows = _load_reglas_store()
        idx = next((i for i, r in enumerate(rows) if int(r.get("id", -1)) == rule_id), None)
        if idx is None:
            raise ValueError("Regla no encontrada.")

        now = datetime.now().isoformat(timespec="seconds")
        rule = dict(rows[idx])
        if action == "toggle":
            rule["estado"] = "Inactiva" if str(rule.get("estado", "")).lower() == "activa" else "Activa"
            rule["ultimo_uso"] = now
            rows[idx] = rule
            _save_reglas_store(rows)
            return rule
        if action == "edit":
            if "condicion" in payload:
                rule["condicion"] = str(payload.get("condicion", "")).strip() or rule.get("condicion")
            if "accion" in payload:
                rule["accion"] = str(payload.get("accion", "")).strip() or rule.get("accion")
            rule["origen"] = "Manual"
            rule["ultimo_uso"] = now
            rows[idx] = rule
            _save_reglas_store(rows)
            return rule
        if action == "duplicate":
            next_id = max((int(r.get("id", 0)) for r in rows), default=0) + 1
            new_rule = dict(rule)
            new_rule["id"] = next_id
            new_rule["nombre"] = f"{rule.get('nombre', 'Regla')} (copia)"
            new_rule["estado"] = "Inactiva"
            new_rule["origen"] = "Manual"
            new_rule["ultimo_uso"] = now
            rows.append(new_rule)
            _save_reglas_store(rows)
            return new_rule
        if action == "delete":
            deleted = rows.pop(idx)
            _save_reglas_store(rows)
            return deleted
        raise ValueError("Acción no soportada.")


# =============================================================================
# Sección 2: Helpers de entrada/salida y transformación de datos
# =============================================================================
def _parse_multipart(content_type: str, body: bytes) -> dict[str, dict[str, object]]:
    match = re.search(r'boundary=(?:"([^"]+)"|([^;]+))', content_type)
    if not match:
        raise ValueError("No se pudo leer el boundary del formulario.")

    boundary = (match.group(1) or match.group(2) or "").strip()
    if not boundary:
      raise ValueError("Boundary invalido.")

    parts: dict[str, dict[str, object]] = {}
    marker = f"--{boundary}".encode("utf-8")

    for chunk in body.split(marker):
        chunk = chunk.strip(b"\r\n")
        if not chunk or chunk == b"--":
            continue
        if b"\r\n\r\n" not in chunk:
            continue

        raw_headers, content = chunk.split(b"\r\n\r\n", 1)
        headers = {}
        for line in raw_headers.decode("utf-8", "replace").split("\r\n"):
            if ":" in line:
                key, value = line.split(":", 1)
                headers[key.lower().strip()] = value.strip()

        disp = headers.get("content-disposition", "")
        name_match = re.search(r'name="([^"]+)"', disp)
        if not name_match:
            continue
        field_name = name_match.group(1)
        file_match = re.search(r'filename="([^"]*)"', disp)
        if file_match and file_match.group(1):
            parts[field_name] = {
                "filename": file_match.group(1),
                "content": content.rstrip(b"\r\n"),
            }
        else:
            parts[field_name] = {
                "value": content.decode("utf-8", "replace").strip(),
            }

    return parts


def _read_upload(parts: dict[str, dict[str, object]]) -> tuple[str, bytes]:
    if "archivo" not in parts:
        raise ValueError("Falta el archivo.")
    item = parts["archivo"]
    filename = os.path.basename(str(item.get("filename", ""))).strip()
    if not filename:
        raise ValueError("Nombre de archivo invalido.")
    data = bytes(item.get("content", b""))
    if not data:
        raise ValueError("El archivo esta vacio.")
    return filename, data


def _extension(filename: str) -> str:
    return Path(filename).suffix.lower()


def _build_preview(filename: str, data: bytes) -> dict:
    ext = _extension(filename)
    valid_ext = ext in {".xlsx", ".csv"}
    prechecks: list[dict[str, str]] = [
        {"label": "Archivo detectado", "state": "Éxito" if filename else "Error", "note": filename},
        {"label": "Extensión válida", "state": "Éxito" if valid_ext else "Error", "note": ".xlsx / .csv" if valid_ext else "Formato no aceptado"},
    ]

    if ext == ".xlsx":
        try:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            sheets = list(wb.sheetnames)
            wb.close()
            prechecks.append({"label": "Hojas detectadas", "state": "Éxito", "note": f"{len(sheets)} hoja(s): {', '.join(sheets[:3])}"})
        except Exception as exc:
            prechecks.append({"label": "Hojas detectadas", "state": "Error", "note": f"No se pudo leer el libro: {exc}"})
            valid_ext = False
    elif ext == ".csv":
        try:
            df = pd.read_csv(io.BytesIO(data))
            prechecks.append({"label": "Hojas detectadas", "state": "No aplica", "note": f"CSV con {len(df)} fila(s) estimadas"})
        except Exception as exc:
            prechecks.append({"label": "Hojas detectadas", "state": "Error", "note": f"No se pudo leer el CSV: {exc}"})
            valid_ext = False
    else:
        prechecks.append({"label": "Hojas detectadas", "state": "Error", "note": "Solo se acepta Excel o CSV"})

    prechecks.append({"label": "Listo para análisis", "state": "Listo" if valid_ext else "Bloqueado", "note": "Puede continuar" if valid_ext else "Revise el archivo"})
    return {"file_name": filename, "extension": ext or "", "size_kb": round(len(data) / 1024, 1), "state": "ready" if valid_ext else "invalid", "prechecks": prechecks}


def _materialize_for_pipeline(filename: str, data: bytes) -> str:
    ext = _extension(filename)
    LOCAL_TMP_ROOT.mkdir(parents=True, exist_ok=True)
    run_dir = LOCAL_TMP_ROOT / f"run_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
    run_dir.mkdir(parents=True, exist_ok=True)

    # Limpieza simple y predecible: conservar solo las 30 corridas mas recientes.
    run_dirs = [p for p in LOCAL_TMP_ROOT.iterdir() if p.is_dir() and p.name.startswith("run_")]
    run_dirs.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    for old_dir in run_dirs[30:]:
        try:
            shutil.rmtree(old_dir, ignore_errors=True)
        except Exception:
            pass

    if ext == ".xlsx":
        path = str(run_dir / filename)
        with open(path, "wb") as f:
            f.write(data)
        return path
    if ext == ".csv":
        csv_path = str(run_dir / (Path(filename).stem + ".csv"))
        xlsx_path = str(run_dir / (Path(filename).stem + ".xlsx"))
        with open(csv_path, "wb") as f:
            f.write(data)
        df = pd.read_csv(csv_path)
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Revisión")
        return xlsx_path
    raise ValueError("Formato no aceptado para el análisis.")


# Helpers de resumen/transformación (pipeline-curated -> UI/API)
def _summarize_pipeline(result: dict, metadata: dict[str, str]) -> dict:
    mapping = result.get("semantic_mapping", [])
    mapped = sum(1 for item in mapping if item.get("match_status") == "mapped")
    ambiguous = sum(1 for item in mapping if item.get("match_status") == "ambiguous")
    unmapped = sum(1 for item in mapping if item.get("match_status") == "unmapped")
    warning_codes = sorted({w.get("code") for w in result.get("warnings_estructurados", []) if isinstance(w, dict) and w.get("code")})
    resumen_corto = (
        f"Mapeo: {mapped} mapeadas, {ambiguous} ambiguas, {unmapped} sin mapear. "
        f"Warnings: {len(warning_codes)} código(s) único(s)."
    )
    mapping_base = []
    for item in mapping:
        if not isinstance(item, dict):
            continue
        source_column = item.get("source_column")
        canonical_field = item.get("canonical_field")
        match_status = item.get("match_status")
        if source_column or canonical_field or match_status:
            mapping_base.append(
                {
                    "source_column": source_column,
                    "canonical_field": canonical_field,
                    "match_status": match_status,
                }
            )

    preview_curado_simple = []
    for col in result.get("normalized_preview", []):
        if not isinstance(col, dict):
            continue
        samples = []
        for sample in col.get("samples", [])[:3]:
            if not isinstance(sample, dict):
                continue
            samples.append(
                {
                    "source_value": sample.get("source_value"),
                    "normalized_value": sample.get("normalized_value"),
                    "confidence": sample.get("confidence"),
                }
            )
        preview_curado_simple.append(
            {
                "source_column": col.get("source_column"),
                "canonical_field": col.get("canonical_field"),
                "value_type": col.get("value_type"),
                "samples": samples,
            }
        )

    normalization_examples = (
        result.get("normalization_summary", {}).get("by_field_and_code_examples", {})
        if isinstance(result.get("normalization_summary", {}), dict)
        else {}
    )
    warning_details = []
    for idx, warning in enumerate(result.get("warnings_estructurados", []), start=1):
        if not isinstance(warning, dict):
            continue
        code = warning.get("code")
        raw_ref = warning.get("raw_ref", {})
        details = warning.get("details", {})
        field_affected = None
        if isinstance(raw_ref, dict):
            field_affected = (
                raw_ref.get("canonical_field")
                or raw_ref.get("field")
                or raw_ref.get("source_column")
                or raw_ref.get("column")
            )

        examples: list[str] = []
        if isinstance(raw_ref, dict):
            for key in ("value", "sample", "samples", "missing", "columns"):
                val = raw_ref.get(key)
                if isinstance(val, list):
                    examples.extend(str(v) for v in val[:3])
                elif val is not None:
                    examples.append(str(val))
        if isinstance(details, dict):
            for key in ("value", "sample", "samples", "missing", "columns"):
                val = details.get(key)
                if isinstance(val, list):
                    examples.extend(str(v) for v in val[:3])
                elif val is not None:
                    examples.append(str(val))

        if isinstance(normalization_examples, dict) and code:
            for field_name, code_map in normalization_examples.items():
                if not isinstance(code_map, dict):
                    continue
                code_examples = code_map.get(code)
                if isinstance(code_examples, list) and code_examples:
                    if not field_affected:
                        field_affected = field_name
                    examples.extend(str(v) for v in code_examples[:3])

        dedup_examples = []
        for e in examples:
            if e and e not in dedup_examples:
                dedup_examples.append(e)

        warning_details.append(
            {
                "observation_id": f"obs_{idx}",
                "code": code,
                "category": warning.get("category"),
                "severity": warning.get("severity"),
                "message": warning.get("message"),
                "field_affected": field_affected,
                "examples": dedup_examples[:5],
            }
        )
    return {
        "file": result.get("file"),
        "sheet": result.get("sheet"),
        "header_row_1based": result.get("header_row_1based"),
        "mapping": {"mapped": mapped, "ambiguous": ambiguous, "unmapped": unmapped},
        "mapping_base": mapping_base,
        "warning_codes": warning_codes,
        "warning_details": warning_details,
        "resumen_corto": resumen_corto,
        "preview_curado_simple": preview_curado_simple,
        "normalization_summary": result.get("normalization_summary", {}),
        "metadata": metadata,
    }


class SmartCounterUIHandler(BaseHTTPRequestHandler):
    server_version = "SmartCounterUI/1.0"

    def log_message(self, format: str, *args) -> None:  # noqa: A003
        return

    # =========================================================================
    # Sección 3: Endpoints API (sin cambios funcionales)
    # =========================================================================
    def _handle_get_api(self, path: str, query: dict[str, list[str]]) -> bool:
        if path == "/health":
            _json_response(self, {"status": "ok"})
            return True
        if path.startswith("/api/revisiones"):
            revisiones = _list_revisiones()
            _json_response(
                self,
                {
                    "revisiones": revisiones,
                    "total": _count_revisiones(),
                },
            )
            return True
        if path == "/api/reglas":
            cliente = (query.get("cliente") or [""])[0]
            reglas = _list_reglas(cliente=cliente or None)
            _json_response(self, {"reglas": reglas, "total": len(reglas)})
            return True
        return False

    def _handle_post_json_api(self, path: str) -> bool:
        if path != "/api/reglas/accion":
            return False
        try:
            length = int(self.headers.get("Content-Length", "0") or "0")
            body = self.rfile.read(length) if length > 0 else b"{}"
            payload = json.loads(body.decode("utf-8") or "{}")
            action = str(payload.get("action", "")).strip().lower()
            rule_id = int(payload.get("rule_id"))
            data = payload.get("payload", {}) if isinstance(payload, dict) else {}
            regla = _apply_regla_action(action, rule_id, payload=data if isinstance(data, dict) else {})
            _json_response(self, {"ok": True, "regla": regla})
        except Exception as exc:
            _json_response(self, {"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
        return True

    def _handle_post_multipart_api(self, path: str, form: dict[str, dict[str, object]]) -> bool:
        if path == "/api/preview":
            filename, data = _read_upload(form)
            _json_response(self, _build_preview(filename, data))
            return True
        if path == "/api/analyze":
            filename, data = _read_upload(form)
            metadata = {
                "cliente": str(form.get("cliente", {}).get("value", "")).strip(),
                "periodo_fiscal": str(form.get("periodo_fiscal", {}).get("value", "")).strip(),
                "tipo_revision": str(form.get("tipo_revision", {}).get("value", "")).strip(),
            }
            materialized_path = _materialize_for_pipeline(filename, data)
            result = run_pipeline_curated(materialized_path)
            summary = _summarize_pipeline(result, metadata)
            summary["file"] = filename
            revision = _persist_revision(summary)
            _json_response(
                self,
                {
                    "estado": "análisis listo",
                    **summary,
                    "revision_guardada": {
                        "revision_id": revision.get("revision_id"),
                        "fecha_hora": revision.get("fecha_hora"),
                    },
                },
            )
            return True
        return False

    def do_GET(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        path = parsed.path
        query = parse_qs(parsed.query)
        if path in {"/", "/index.html"}:
            _html_response(self, build_html())
            return
        if self._handle_get_api(path, query):
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        path = parsed.path
        if self._handle_post_json_api(path):
            return

        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            self.send_error(HTTPStatus.BAD_REQUEST, "Se esperaba multipart/form-data.")
            return
        length = int(self.headers.get("Content-Length", "0") or "0")
        body = self.rfile.read(length)
        form = _parse_multipart(content_type, body)
        try:
            if self._handle_post_multipart_api(path, form):
                return
            self.send_error(HTTPStatus.NOT_FOUND)
        except Exception as exc:
            _json_response(self, {"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)


def build_html() -> str:
    return "".join(HTML_PARTS)


# =============================================================================
# Arranque local de servidor UI
# =============================================================================
def serve(host: str, port: int) -> None:
    server = ThreadingHTTPServer((host, port), SmartCounterUIHandler)
    url = f"http://{host}:{port}"
    print(f"SmartCounter UI en {url}")
    threading.Timer(0.75, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Pantalla local SmartCounter - Ingreso de archivo")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8501)
    args = parser.parse_args()
    serve(args.host, args.port)
    return 0


HTML_PARTS.extend([
    r"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>SmartCounter - Ingreso de Archivo</title>
  <style>
    :root {
      --bg:#f4f6fa; --sidebar:#eef2f6; --panel:#fff; --panel-soft:#f8fafc; --border:#d9e1ea;
      --text:#1d2b3a; --muted:#66778b; --primary:#5d6c84; --primary-dark:#49596f;
      --accent:#67d3b7; --check:#4fb6a0; --danger:#c96c6c; --shadow:0 20px 40px rgba(23,39,63,.08);
    }
    *{box-sizing:border-box} body{margin:0;font-family:"Segoe UI","Aptos",sans-serif;background:var(--bg);color:var(--text)}
    .shell{display:grid;grid-template-columns:300px 1fr;min-height:100vh}
    .sidebar{background:var(--sidebar);border-right:1px solid rgba(105,122,142,.15);padding:28px 20px 22px;display:flex;flex-direction:column;gap:26px}
    .brand h1{margin:0;font-size:28px;line-height:1;font-weight:700;letter-spacing:-.02em}.brand .sub{margin-top:8px;color:#67788f;font-size:12px;letter-spacing:.16em;text-transform:uppercase}
    .nav{display:grid;gap:10px;margin-top:8px}.nav a{text-decoration:none;color:#5f7086;display:flex;align-items:center;gap:12px;padding:14px 12px;border-radius:12px;font-size:18px}.nav a.active{background:rgba(93,108,132,.08);color:var(--text);font-weight:600}.nav .icon{width:24px;height:24px;border-radius:6px;background:rgba(93,108,132,.22);flex:none}
    .sidebar-cta{margin-top:auto;padding-top:16px;border-top:1px solid rgba(105,122,142,.16)}.sidebar-btn{width:100%;background:var(--primary);color:#fff;border:0;border-radius:3px;padding:16px;font-size:16px;font-weight:700;box-shadow:var(--shadow);cursor:pointer}
    .sidebar-foot{display:grid;gap:6px;margin-top:18px}.sidebar-foot a{color:#5d6d82;text-decoration:none;font-weight:700;letter-spacing:.08em;text-transform:uppercase;font-size:13px;padding:10px 0;display:flex;gap:10px;align-items:center}
    .main{padding:18px 28px 28px}.topbar{display:flex;align-items:center;justify-content:space-between;padding:10px 6px 18px}.crumbs{color:#263445;font-size:15px;display:flex;gap:10px;align-items:center}
    .userbox{display:flex;align-items:center;gap:14px;color:#334155}.avatar{width:38px;height:38px;border-radius:50%;background:linear-gradient(135deg,#d9e3ef,#9fb1c6);border:1px solid rgba(86,103,124,.18)}
    .page-head{display:flex;align-items:center;gap:14px;flex-wrap:wrap;padding:8px 6px 22px}.page-title{margin:0;font-size:60px;line-height:1;letter-spacing:-.04em}.pill{background:#e8edf3;color:#4b5d72;border:1px solid #d2dae3;border-radius:3px;padding:8px 12px;font-size:13px;letter-spacing:.06em;text-transform:uppercase;font-weight:700}
    .subtitle{padding:0 6px 20px;color:#526273;font-size:20px}.grid{display:grid;grid-template-columns:1fr 360px;gap:28px;align-items:start;margin-top:4px}
    .panel{background:var(--panel);border:1px solid rgba(214,222,231,.8);box-shadow:var(--shadow)} .drop-zone{min-height:620px;padding:88px 48px 36px;position:relative}.drop-zone.dragover{outline:3px dashed rgba(79,182,160,.55);outline-offset:-12px;background:#fbfdff}
    .doc-fade{position:absolute;right:30px;top:24px;width:54px;height:54px;opacity:.18;background:linear-gradient(180deg,#dfe8f2,#eef3f8);border-radius:4px}
    .uploader{display:grid;place-items:center;gap:22px;text-align:center}.upload-badge{width:84px;height:84px;border-radius:16px;background:#ecf1f7;display:grid;place-items:center;color:#5b6b7f;font-size:34px;font-weight:900}
    .uploader h2{margin:0;font-size:28px;letter-spacing:-.03em}.uploader p{max-width:540px;margin:0;color:#5c6b7e;font-size:18px;line-height:1.45}
    .file-card{width:min(560px,100%);display:flex;align-items:center;justify-content:space-between;gap:16px;border-radius:6px;background:#eef3f7;border:1px solid #d7e1ea;padding:18px;margin-top:14px}
    .file-left{display:flex;align-items:center;gap:16px;min-width:0}.file-icon{width:48px;height:58px;background:#fff;border:1px solid #dbe3ec;border-radius:2px;display:grid;place-items:center;color:#1fa56b;font-weight:800;flex:none}.file-name{font-size:18px;font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.file-meta{color:#566577;font-size:14px;margin-top:4px}
    .icon-btn{border:0;background:transparent;color:#516173;font-size:20px;cursor:pointer;padding:6px}.ghost-btn{margin-top:30px;background:transparent;border:2px dashed #b8c2cd;color:#43546a;padding:18px 34px;font-size:18px;font-weight:700;border-radius:4px;cursor:pointer}
    .right-stack{display:grid;gap:18px}.checks{background:#67758b;color:#fff;padding:30px 28px;border-radius:3px;box-shadow:var(--shadow)}.checks h3{margin:0 0 24px;font-size:22px}
    .check-row{display:flex;align-items:center;justify-content:space-between;gap:16px;padding:11px 0;border-bottom:1px solid rgba(255,255,255,.08);color:#edf3fa;font-size:18px}.check-row:last-child{border-bottom:0}.check-state{color:#76efd0;font-size:12px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;min-width:68px;text-align:right}
    .info{background:#f7fafc;border-left:6px solid #c7d0db;padding:36px 30px;color:#405165}.info h3{margin:0 0 16px;font-size:22px;color:#263346}.info p{margin:0 0 18px;line-height:1.58;font-size:16px;color:#576779}.info a{color:#405574;font-weight:700;text-decoration:none}
    .meta-card{margin-top:22px;padding:30px 34px;display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:26px;align-items:end}.field label{display:block;font-size:13px;letter-spacing:.2em;text-transform:uppercase;color:#505d6d;margin-bottom:18px;font-weight:700}.field input,.field select{width:100%;border:0;border-bottom:2px solid #d9e2ea;background:transparent;color:#152433;font-size:18px;padding:6px 0 10px;outline:none}
    .field input:focus,.field select:focus{border-bottom-color:#64758b}.bottom-bar{display:flex;align-items:center;justify-content:space-between;gap:18px;margin-top:26px;padding:16px 6px 0;border-top:1px solid rgba(204,213,224,.8)}
    .bottom-actions{display:flex;gap:22px;flex-wrap:wrap;color:#3e4d60;font-weight:700;align-items:center}.bottom-actions button{border:0;background:transparent;color:inherit;font:inherit;cursor:pointer;padding:10px 0}
    .primary-btn{background:#4f5f77;color:#fff;border:0;padding:18px 34px;font-size:18px;font-weight:800;border-radius:3px;cursor:pointer;min-width:260px;box-shadow:var(--shadow)}.primary-btn:disabled{opacity:.45;cursor:not-allowed}
    .status-line{padding:14px 6px 0;color:#526273;font-size:14px;min-height:22px}.status-line.error{color:var(--danger);font-weight:700}
    .result{margin-top:18px;padding:22px 24px;background:#f8fbfd;border:1px solid #dce4ec}.result h4{margin:0 0 12px;font-size:18px}.result pre{margin:0;white-space:pre-wrap;word-break:break-word;color:#425163;font-size:14px;line-height:1.55}
    .screen2{display:none}
    .screen2.active{display:block}
    .screen2-head{display:flex;align-items:center;justify-content:space-between;gap:16px;margin:6px 6px 14px}
    .screen2-badge{background:#1f2d3d;color:#fff;padding:8px 12px;border-radius:4px;font-size:12px;letter-spacing:.08em;text-transform:uppercase;font-weight:700}
    .screen2-title{margin:0;font-size:32px;letter-spacing:-.02em}
    .screen2-sub{margin:6px 0 0;color:#53657a;font-size:16px}
    .screen2-actions{display:flex;gap:10px}
    .secondary-btn{border:1px solid #c9d4e0;background:#fff;color:#364a62;padding:10px 14px;font-size:14px;font-weight:700;border-radius:4px;cursor:pointer}
    .screen2 .result{margin-top:0;padding:26px 28px;background:#f4f8fc;border:1px solid #cfdae7}
    .screen2 .result h4{font-size:22px}
    .screen2 .result pre{font-size:15px;line-height:1.6}
    .obs-panel{margin-top:16px;padding:18px 20px;background:#fff;border:1px solid #d6e0eb}
    .obs-panel h4{margin:0 0 10px;font-size:18px}
    .obs-list{display:flex;flex-wrap:wrap;gap:10px}
    .obs-btn{border:1px solid #cbd6e4;background:#f7fbff;color:#32465f;padding:8px 12px;font-size:13px;font-weight:700;border-radius:4px;cursor:pointer}
    .screen3{display:none}
    .screen3.active{display:block}
    .screen3 .result{margin-top:0;padding:26px 28px;background:#fff;border:1px solid #cfdae7}
    .detail-label{font-size:12px;letter-spacing:.08em;text-transform:uppercase;color:#5e6f84;font-weight:700}
    .detail-value{margin:4px 0 14px;font-size:16px;color:#233447}
    .detail-actions{display:flex;gap:10px;flex-wrap:wrap;margin-top:10px}
    .primary-compact{border:0;background:#4f5f77;color:#fff;padding:10px 14px;font-size:14px;font-weight:700;border-radius:4px;cursor:pointer}
    .manual-note{margin-top:10px;color:#49627f;font-size:14px}
    .screen4{display:none}
    .screen4.active{display:block}
    .rules-layout{display:grid;grid-template-columns:1fr 360px;gap:16px}
    .rules-list{padding:18px 20px;background:#fff;border:1px solid #d6e0eb}
    .rules-row{display:grid;grid-template-columns:2fr 1fr 1fr 1fr 1fr auto;gap:10px;align-items:center;padding:10px 0;border-bottom:1px solid #e6edf5;font-size:13px}
    .rules-row.head{font-size:10px;text-transform:uppercase;letter-spacing:.08em;color:#617388;font-weight:700}
    .rules-row:last-child{border-bottom:0}
    .rules-detail{padding:18px 20px;background:#f8fbff;border:1px solid #d6e0eb}
    .rules-detail .detail-value{font-size:14px}
    .rules-empty{font-size:14px;color:#5f7186;padding:12px 0}
    .screen5{display:none}
    .screen5.active{display:block}
    .history-wrap{display:grid;grid-template-columns:1fr;gap:16px}
    .history-list{padding:18px 20px;background:#fff;border:1px solid #d6e0eb}
    .history-item{display:flex;align-items:flex-start;justify-content:space-between;gap:14px;padding:12px 0;border-bottom:1px solid #e5edf5}
    .history-item:last-child{border-bottom:0}
    .history-meta{font-size:13px;color:#55697f}
    .history-title{font-size:16px;font-weight:700;color:#233447}
    .history-empty{font-size:15px;color:#5f7186}
    .history-detail{padding:22px 24px;background:#f8fbff;border:1px solid #d6e0eb}
    .history-detail pre{margin:0;white-space:pre-wrap;word-break:break-word;color:#425163;font-size:14px;line-height:1.55}
    .screen6{display:none}
    .screen6.active{display:block}
    .close-wrap{display:grid;grid-template-columns:1fr;gap:16px}
    .close-summary{padding:22px 24px;background:#fff;border:1px solid #d6e0eb}
    .close-grid{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:12px}
    .close-cell{background:#f3f8fd;padding:12px 14px}
    .close-k{font-size:10px;letter-spacing:.08em;text-transform:uppercase;color:#627388;font-weight:700}
    .close-v{margin-top:4px;font-size:14px;color:#24374c;font-weight:700}
    .status-chip{display:inline-block;padding:6px 10px;font-size:11px;font-weight:700;letter-spacing:.06em;text-transform:uppercase}
    .status-ok{background:#dff4ea;color:#1f6d4f}
    .status-warn{background:#fff2dd;color:#895f12}
    .status-critical{background:#fbe3e3;color:#8e3737}
    .close-notes{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px}
    .close-notes textarea{width:100%;min-height:90px;border:0;border-bottom:2px solid #cad7e5;background:#eef4fa;padding:10px 12px;color:#2b3e52;resize:vertical}
    .close-actions{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}
    .action-btn{display:flex;align-items:center;justify-content:space-between;gap:12px;padding:14px 16px;border:1px solid #d2deea;background:#fff;color:#2a3e55;font-weight:700;cursor:pointer}
    .action-btn.primary{background:#4f5f77;color:#fff;border-color:#4f5f77}
    @media (max-width:1200px){.grid{grid-template-columns:1fr}.meta-card{grid-template-columns:1fr}.rules-layout{grid-template-columns:1fr}.close-grid{grid-template-columns:1fr 1fr}.close-notes{grid-template-columns:1fr}}
    dialog{border:1px solid #d1dbe5;border-radius:8px;padding:0;max-width:560px;width:calc(100vw - 32px);box-shadow:var(--shadow)}dialog::backdrop{background:rgba(18,28,43,.36)}
    .modal{padding:22px 24px 18px;background:#fff}.modal h4{margin:0 0 10px;font-size:22px}.modal p,.modal li{color:#526273;line-height:1.5}.modal-actions{display:flex;justify-content:flex-end;gap:12px;padding-top:16px}.modal-actions button{border:0;background:#4f5f77;color:#fff;padding:10px 16px;border-radius:4px;cursor:pointer}
    @media (max-width:1200px){.grid{grid-template-columns:1fr}.meta-card{grid-template-columns:1fr}}
    @media (max-width:900px){.shell{grid-template-columns:1fr}.sidebar{border-right:0;border-bottom:1px solid rgba(105,122,142,.15)}.page-title{font-size:42px}}
  </style>
</head>
<body>
  <div class="shell">
    <aside class="sidebar">
      <div class="brand"><h1>SmartCounter</h1><div class="sub">Argentum Ledger</div></div>
      <nav class="nav">
        <a class="active" href="javascript:void(0)"><span class="icon"></span>Revisiones</a>
        <a href="javascript:void(0)"><span class="icon"></span>Clientes</a>
        <a href="javascript:void(0)" id="rulesNavBtn"><span class="icon"></span>Reglas</a>
        <a href="javascript:void(0)" id="historyNavBtn"><span class="icon"></span>Historial</a>
      </nav>
      <div class="sidebar-cta">
        <button class="sidebar-btn" type="button" id="newFileBtn">Nuevo expediente</button>
        <div class="sidebar-foot"><a href="javascript:void(0)">Configuración</a><a href="javascript:void(0)">Ayuda</a></div>
      </div>
    </aside>
    <main class="main">
      <div class="topbar">
        <div class="crumbs">Inicio <span>›</span> Ingreso de archivo</div>
        <div class="userbox"><div style="font-size:24px;color:#74859a;">☻</div><div><div style="font-weight:700;">Usuario contador</div><div style="font-size:13px;color:#6c7b8e;">Matrícula 44.231</div></div><div class="avatar"></div></div>
      </div>
      <div class="page-head"><h2 class="page-title" id="pageTitle">Sin archivo cargado</h2><span class="pill" id="statusPill">Esperando selección</span></div>
      <div class="subtitle" id="pageSubtitle">Seleccione un archivo Excel o CSV para iniciar la revisión.</div>
      <!-- Sección 4: Render / template Pantalla 1 -->
      <section class="grid" id="screen1Area">
        <div>
          <div class="panel drop-zone" id="dropZone">
            <div class="doc-fade"></div>
            <div class="uploader">
              <div class="upload-badge">↥</div>
              <h2>Subir archivo de revisión</h2>
              <p>Arrastre un archivo Excel o CSV para iniciar la revisión automatizada del período seleccionado.</p>
              <div id="fileCard" class="file-card" style="display:none;"><div class="file-left"><div class="file-icon">X</div><div style="min-width:0;"><div class="file-name" id="fileName"></div><div class="file-meta" id="fileMeta"></div></div></div><button class="icon-btn" type="button" id="removeFileBtn" title="Quitar archivo">🗑</button></div>
              <button class="ghost-btn" type="button" id="pickFileBtn">Seleccionar otro archivo</button>
              <input type="file" id="fileInput" accept=".xlsx,.csv" hidden />
            </div>
          </div>
          <div class="panel meta-card">
            <div class="field"><label>Cliente</label><input id="cliente" type="text" placeholder="Nombre del cliente" /></div>
            <div class="field"><label>Período fiscal</label><input id="periodoFiscal" type="text" placeholder="03/2026" /></div>
            <div class="field"><label>Tipo de revisión</label><select id="tipoRevision"><option value="IVA compras">IVA compras</option><option value="IVA ventas">IVA ventas</option><option value="Cobranzas / pagos">Cobranzas / pagos</option><option value="Mayor auxiliar">Mayor auxiliar</option><option value="Balance">Balance</option><option value="Otro">Otro</option></select></div>
          </div>
          <div class="bottom-bar">
            <div class="bottom-actions"><button type="button" id="clearBtn">Limpiar selección</button><button type="button" id="exampleBtn">Ver ejemplo de archivo</button></div>
            <button class="primary-btn" type="button" id="analyzeBtn" disabled>Iniciar análisis ▶</button>
          </div>
          <div class="status-line" id="statusLine"></div>
        </div>
        <div class="right-stack">
          <section class="checks panel" id="checksPanel"><h3>Chequeos previos</h3><div id="checksList"></div></section>
          <section class="info panel"><h3>Ayuda operativa</h3><p>Esta fase valida la integridad mínima del archivo antes de ejecutar las reglas de interpretación. El motor queda listo para conectarse luego con el pipeline-curated ya confirmado.</p><p><a href="javascript:void(0)" id="downloadGuide">Ver guía de formatos</a></p></section>
        </div>
      </section>
      <!-- Sección 5: Render / template Pantalla 2 -->
      <section class="screen2" id="screen2Area">
        <div class="screen2-head">
          <div>
            <span class="screen2-badge">Resultado curado</span>
            <h3 class="screen2-title">Vista de análisis</h3>
            <p class="screen2-sub">Salida del pipeline-curated para revisión humana.</p>
          </div>
          <div class="screen2-actions">
            <button type="button" class="secondary-btn" id="backToCargaBtn">Volver a carga</button>
            <button type="button" class="secondary-btn" id="openCloseBtn">Cierre / exportación</button>
          </div>
        </div>
        <div class="result" id="resultBox" style="display:none;"><h4>Resultado curado</h4><pre id="resultText"></pre></div>
        <div class="obs-panel" id="observationPanel" style="display:none;">
          <h4>Observaciones detectadas</h4>
          <div class="obs-list" id="observationList"></div>
        </div>
      </section>
      <!-- Sección 6: Render / template Pantalla 3 -->
      <section class="screen3" id="screen3Area">
        <div class="screen2-head">
          <div>
            <span class="screen2-badge">Detalle de observación</span>
            <h3 class="screen2-title">Regla aplicada</h3>
            <p class="screen2-sub">Vista simple para revisar una observación puntual.</p>
          </div>
        </div>
        <div class="result">
          <h4>Detalle</h4>
          <div class="detail-label">Código</div><div class="detail-value" id="detailCode">No disponible</div>
          <div class="detail-label">Categoría / tipo</div><div class="detail-value" id="detailCategory">No disponible</div>
          <div class="detail-label">Severidad</div><div class="detail-value" id="detailSeverity">No disponible</div>
          <div class="detail-label">Mensaje</div><div class="detail-value" id="detailMessage">No disponible</div>
          <div class="detail-label">Campo afectado</div><div class="detail-value" id="detailField">No disponible</div>
          <div class="detail-label">Ejemplos / valores observados</div><div class="detail-value" id="detailExamples">No disponible</div>
          <div class="detail-actions">
            <button type="button" class="primary-compact" id="markManualBtn">Marcar para revisión manual</button>
            <button type="button" class="secondary-btn" id="backToResultBtn">Volver al resultado curado</button>
          </div>
          <div class="manual-note" id="manualReviewNote"></div>
        </div>
      </section>
      <!-- Sección 7: Render / template Pantalla 4 -->
      <section class="screen4" id="screen4Area">
        <div class="screen2-head">
          <div>
            <span class="screen2-badge">Reglas del cliente</span>
            <h3 class="screen2-title">Motor de reglas operativas</h3>
            <p class="screen2-sub">Listado y detalle simple para activación y mantenimiento.</p>
          </div>
          <div class="screen2-actions">
            <button type="button" class="secondary-btn" id="rulesBackToCargaBtn">Volver a carga</button>
          </div>
        </div>
        <div class="rules-layout">
          <div class="rules-list panel">
            <div class="rules-row head">
              <div>Nombre</div><div>Categoría</div><div>Origen</div><div>Estado</div><div>Último uso</div><div></div>
            </div>
            <div id="rulesList"></div>
          </div>
          <div class="rules-detail panel">
            <h4 style="margin:0 0 12px;font-size:18px;">Detalle de regla</h4>
            <div class="detail-label">Cliente</div><div class="detail-value" id="ruleDetailCliente">No disponible</div>
            <div class="detail-label">Tipo</div><div class="detail-value" id="ruleDetailTipo">No disponible</div>
            <div class="detail-label">Condición</div><div class="detail-value" id="ruleDetailCondicion">No disponible</div>
            <div class="detail-label">Acción</div><div class="detail-value" id="ruleDetailAccion">No disponible</div>
            <div class="detail-label">Ejemplos</div><div class="detail-value" id="ruleDetailEjemplos">No disponible</div>
            <div class="detail-label">Impacto esperado</div><div class="detail-value" id="ruleDetailImpacto">No disponible</div>
            <div class="detail-actions">
              <button type="button" class="secondary-btn" id="ruleToggleBtn">Activar / desactivar</button>
              <button type="button" class="secondary-btn" id="ruleEditBtn">Editar</button>
              <button type="button" class="secondary-btn" id="ruleDuplicateBtn">Duplicar</button>
              <button type="button" class="secondary-btn" id="ruleDeleteBtn">Eliminar</button>
            </div>
          </div>
        </div>
      </section>
      <!-- Sección 8: Render / template Pantalla 5 -->
      <section class="screen5" id="screen5Area">
        <div class="screen2-head">
          <div>
            <span class="screen2-badge">Historial</span>
            <h3 class="screen2-title">Revisiones guardadas</h3>
            <p class="screen2-sub">Listado local de revisiones procesadas.</p>
          </div>
          <div class="screen2-actions">
            <button type="button" class="secondary-btn" id="historyBackToCargaBtn">Volver a carga</button>
          </div>
        </div>
        <div class="history-wrap">
          <div class="history-list panel">
            <h4 style="margin:0 0 12px;font-size:20px;">Listado</h4>
            <div id="historyList"></div>
          </div>
          <div class="history-detail panel" id="historyDetail" style="display:none;">
            <h4 style="margin:0 0 12px;font-size:20px;">Revisión seleccionada</h4>
            <pre id="historyDetailText"></pre>
            <div class="detail-actions" style="margin-top:14px;">
              <button type="button" class="secondary-btn" id="historyBackToListBtn">Volver al listado</button>
              <button type="button" class="secondary-btn" id="historyOpenInResultBtn">Ver como revisión seleccionada</button>
              <button type="button" class="secondary-btn" id="historyOpenCloseBtn">Ir a cierre</button>
            </div>
          </div>
        </div>
      </section>
      <!-- Sección 9: Render / template Pantalla 6 -->
      <section class="screen6" id="screen6Area">
        <div class="screen2-head">
          <div>
            <span class="screen2-badge">Cierre de revisión</span>
            <h3 class="screen2-title">Salida y trazabilidad</h3>
            <p class="screen2-sub">Resumen final y acciones de exportación.</p>
          </div>
        </div>
        <div class="close-wrap">
          <div class="close-summary panel">
            <h4 style="margin:0 0 12px;font-size:20px;">Resumen final</h4>
            <div class="close-grid">
              <div class="close-cell"><div class="close-k">Archivo</div><div class="close-v" id="closeFile">No disponible</div></div>
              <div class="close-cell"><div class="close-k">Hoja detectada</div><div class="close-v" id="closeSheet">No disponible</div></div>
              <div class="close-cell"><div class="close-k">Fila header</div><div class="close-v" id="closeHeader">No disponible</div></div>
              <div class="close-cell"><div class="close-k">warning_codes</div><div class="close-v" id="closeWarnings">No disponible</div></div>
              <div class="close-cell"><div class="close-k">Resumen corto</div><div class="close-v" id="closeSummary">No disponible</div></div>
            </div>
            <div style="margin-top:14px;">
              <div class="close-k">Estado final sugerido</div>
              <div id="closeStatusChip" class="status-chip status-warn">Con observaciones</div>
            </div>
          </div>
          <div class="close-summary panel">
            <h4 style="margin:0 0 12px;font-size:20px;">Notas de cierre</h4>
            <div class="close-notes">
              <div><div class="close-k">Observación final</div><textarea id="closeFinalNote" placeholder="Ingrese observación final..."></textarea></div>
              <div><div class="close-k">Comentario de trazabilidad</div><textarea id="closeTraceNote" placeholder="Ingrese comentario de trazabilidad..."></textarea></div>
            </div>
          </div>
          <div class="close-summary panel">
            <h4 style="margin:0 0 12px;font-size:20px;">Acciones de salida</h4>
            <div class="close-actions">
              <button type="button" class="action-btn primary" id="downloadCuratedBtn"><span>Descargar salida curada</span><span>⬇</span></button>
              <button type="button" class="action-btn" id="downloadReviewSummaryBtn"><span>Descargar resumen de revisión</span><span>⬇</span></button>
              <button type="button" class="action-btn" id="closeBackToHistoryBtn"><span>Volver al historial</span><span>↩</span></button>
              <button type="button" class="action-btn" id="closeStartNewBtn"><span>Iniciar nueva revisión</span><span>↻</span></button>
            </div>
            <div class="history-meta" style="margin-top:10px;">La exportación es simple y local en esta versión.</div>
          </div>
        </div>
      </section>
    </main>
  </div>
  <dialog id="exampleDialog"><div class="modal"><h4>Ejemplo de archivo válido</h4><p>Un archivo adecuado para esta pantalla suele tener:</p><ul><li>una hoja con tabla útil y encabezado claro;</li><li>columnas de cliente, fecha, monto o saldo;</li><li>formatos legibles para revisión humana;</li><li>sin promesas de automatización total.</li></ul><div class="modal-actions"><button type="button" id="closeDialog">Cerrar</button></div></div></dialog>
""",
])


if __name__ == "__main__":
    raise SystemExit(main())

HTML_PARTS.extend([
    r"""
  <!-- Sección 10: Navegación / estado UI -->
  <script>
    const state = {
      file: null,
      preview: null,
      valid: false,
      analysis: null,
      selectedObservationId: null,
      manualReview: {},
      rules: [],
      selectedRule: null,
      history: [],
      selectedRevision: null,
      closeContext: null,
    };
    const fileInput = document.getElementById('fileInput');
    const pickFileBtn = document.getElementById('pickFileBtn');
    const newFileBtn = document.getElementById('newFileBtn');
    const removeFileBtn = document.getElementById('removeFileBtn');
    const clearBtn = document.getElementById('clearBtn');
    const exampleBtn = document.getElementById('exampleBtn');
    const analyzeBtn = document.getElementById('analyzeBtn');
    const statusLine = document.getElementById('statusLine');
    const resultBox = document.getElementById('resultBox');
    const resultText = document.getElementById('resultText');
    const screen1Area = document.getElementById('screen1Area');
    const screen2Area = document.getElementById('screen2Area');
    const screen3Area = document.getElementById('screen3Area');
    const screen4Area = document.getElementById('screen4Area');
    const screen5Area = document.getElementById('screen5Area');
    const screen6Area = document.getElementById('screen6Area');
    const backToCargaBtn = document.getElementById('backToCargaBtn');
    const openCloseBtn = document.getElementById('openCloseBtn');
    const backToResultBtn = document.getElementById('backToResultBtn');
    const markManualBtn = document.getElementById('markManualBtn');
    const manualReviewNote = document.getElementById('manualReviewNote');
    const observationPanel = document.getElementById('observationPanel');
    const observationList = document.getElementById('observationList');
    const detailCode = document.getElementById('detailCode');
    const detailCategory = document.getElementById('detailCategory');
    const detailSeverity = document.getElementById('detailSeverity');
    const detailMessage = document.getElementById('detailMessage');
    const detailField = document.getElementById('detailField');
    const detailExamples = document.getElementById('detailExamples');
    const pageTitle = document.getElementById('pageTitle');
    const pageSubtitle = document.getElementById('pageSubtitle');
    const statusPill = document.getElementById('statusPill');
    const fileCard = document.getElementById('fileCard');
    const fileName = document.getElementById('fileName');
    const fileMeta = document.getElementById('fileMeta');
    const checksList = document.getElementById('checksList');
    const dropZone = document.getElementById('dropZone');
    const cliente = document.getElementById('cliente');
    const periodoFiscal = document.getElementById('periodoFiscal');
    const tipoRevision = document.getElementById('tipoRevision');
    const dialog = document.getElementById('exampleDialog');
    const closeDialog = document.getElementById('closeDialog');
    const downloadGuide = document.getElementById('downloadGuide');
    const historyNavBtn = document.getElementById('historyNavBtn');
    const rulesNavBtn = document.getElementById('rulesNavBtn');
    const rulesBackToCargaBtn = document.getElementById('rulesBackToCargaBtn');
    const rulesList = document.getElementById('rulesList');
    const ruleDetailCliente = document.getElementById('ruleDetailCliente');
    const ruleDetailTipo = document.getElementById('ruleDetailTipo');
    const ruleDetailCondicion = document.getElementById('ruleDetailCondicion');
    const ruleDetailAccion = document.getElementById('ruleDetailAccion');
    const ruleDetailEjemplos = document.getElementById('ruleDetailEjemplos');
    const ruleDetailImpacto = document.getElementById('ruleDetailImpacto');
    const ruleToggleBtn = document.getElementById('ruleToggleBtn');
    const ruleEditBtn = document.getElementById('ruleEditBtn');
    const ruleDuplicateBtn = document.getElementById('ruleDuplicateBtn');
    const ruleDeleteBtn = document.getElementById('ruleDeleteBtn');
    const historyBackToCargaBtn = document.getElementById('historyBackToCargaBtn');
    const historyBackToListBtn = document.getElementById('historyBackToListBtn');
    const historyOpenInResultBtn = document.getElementById('historyOpenInResultBtn');
    const historyOpenCloseBtn = document.getElementById('historyOpenCloseBtn');
    const historyList = document.getElementById('historyList');
    const historyDetail = document.getElementById('historyDetail');
    const historyDetailText = document.getElementById('historyDetailText');
    const closeFile = document.getElementById('closeFile');
    const closeSheet = document.getElementById('closeSheet');
    const closeHeader = document.getElementById('closeHeader');
    const closeWarnings = document.getElementById('closeWarnings');
    const closeSummary = document.getElementById('closeSummary');
    const closeStatusChip = document.getElementById('closeStatusChip');
    const closeFinalNote = document.getElementById('closeFinalNote');
    const closeTraceNote = document.getElementById('closeTraceNote');
    const downloadCuratedBtn = document.getElementById('downloadCuratedBtn');
    const downloadReviewSummaryBtn = document.getElementById('downloadReviewSummaryBtn');
    const closeBackToHistoryBtn = document.getElementById('closeBackToHistoryBtn');
    const closeStartNewBtn = document.getElementById('closeStartNewBtn');

    function showScreen1() {
      screen1Area.style.display = 'grid';
      screen2Area.classList.remove('active');
      screen3Area.classList.remove('active');
      screen4Area.classList.remove('active');
      screen5Area.classList.remove('active');
      screen6Area.classList.remove('active');
      if (state.file) {
        pageTitle.textContent = state.file.name;
        pageSubtitle.textContent = 'Cliente y período fiscal listos para confirmar.';
        statusPill.textContent = state.valid ? 'Pendiente de análisis' : 'Archivo inválido';
        statusPill.style.background = state.valid ? '#e8edf3' : '#f8e2e2';
        statusPill.style.color = state.valid ? '#4b5d72' : '#8f4c4c';
      } else {
        pageTitle.textContent = 'Sin archivo cargado';
        pageSubtitle.textContent = 'Seleccione un archivo Excel o CSV para iniciar la revisión.';
        statusPill.textContent = 'Esperando selección';
        statusPill.style.background = '#e8edf3';
        statusPill.style.color = '#4b5d72';
      }
    }

    function showScreen2() {
      screen1Area.style.display = 'none';
      screen2Area.classList.add('active');
      screen3Area.classList.remove('active');
      screen4Area.classList.remove('active');
      screen5Area.classList.remove('active');
      screen6Area.classList.remove('active');
    }

    function showScreen3() {
      screen1Area.style.display = 'none';
      screen2Area.classList.remove('active');
      screen3Area.classList.add('active');
      screen4Area.classList.remove('active');
      screen5Area.classList.remove('active');
      screen6Area.classList.remove('active');
    }

    function showScreen4() {
      screen1Area.style.display = 'none';
      screen2Area.classList.remove('active');
      screen3Area.classList.remove('active');
      screen4Area.classList.add('active');
      screen5Area.classList.remove('active');
      screen6Area.classList.remove('active');
      pageTitle.textContent = 'Reglas del cliente';
      pageSubtitle.textContent = 'Gestione reglas activas y sugeridas para el cliente.';
      statusPill.textContent = 'Reglas';
      statusPill.style.background = '#eef2f7';
      statusPill.style.color = '#30445d';
    }

    function showScreen5() {
      screen1Area.style.display = 'none';
      screen2Area.classList.remove('active');
      screen3Area.classList.remove('active');
      screen4Area.classList.remove('active');
      screen5Area.classList.add('active');
      screen6Area.classList.remove('active');
      pageTitle.textContent = 'Historial de revisiones';
      pageSubtitle.textContent = 'Listado local de revisiones procesadas.';
      statusPill.textContent = 'Historial';
      statusPill.style.background = '#eef2f7';
      statusPill.style.color = '#30445d';
    }

    function showScreen6() {
      screen1Area.style.display = 'none';
      screen2Area.classList.remove('active');
      screen3Area.classList.remove('active');
      screen4Area.classList.remove('active');
      screen5Area.classList.remove('active');
      screen6Area.classList.add('active');
      pageTitle.textContent = 'Cierre / exportación';
      pageSubtitle.textContent = 'Valide el estado final y descargue los artefactos de salida.';
      statusPill.textContent = 'Cierre';
      statusPill.style.background = '#e9f1fb';
      statusPill.style.color = '#2f4760';
    }

    function setStatus(text, kind='') {
      statusLine.textContent = text || '';
      statusLine.className = 'status-line' + (kind ? ' ' + kind : '');
    }

    function renderChecks(prechecks) {
      checksList.innerHTML = '';
      prechecks.forEach(item => {
        const row = document.createElement('div');
        row.className = 'check-row';
        row.innerHTML = `<div>${item.label}</div><div style="flex:1;text-align:right;color:#dce8ff;">${item.note}</div><div class="check-state">${item.state}</div>`;
        checksList.appendChild(row);
      });
    }

    function renderCuratedResult(data) {
      state.analysis = data;
      const lines = [];
      const warningCodes = Array.isArray(data.warning_codes) ? data.warning_codes : [];
      const mappingBase = Array.isArray(data.mapping_base) ? data.mapping_base : [];
      const previewCurado = Array.isArray(data.preview_curado_simple) ? data.preview_curado_simple : [];

      lines.push('Contexto del archivo');
      lines.push(`- Archivo: ${data.file || 'No disponible'}`);
      if (data.sheet) lines.push(`- Hoja detectada: ${data.sheet}`);
      if (data.header_row_1based) lines.push(`- Fila header: ${data.header_row_1based}`);

      lines.push('');
      lines.push('warning_codes');
      lines.push(warningCodes.length ? `- ${warningCodes.join(', ')}` : '- Sin warning_codes');

      if (data.mapping && typeof data.mapping === 'object') {
        lines.push('');
        lines.push('Resumen de mapping');
        lines.push(`- mapped: ${data.mapping.mapped ?? 0}`);
        lines.push(`- ambiguous: ${data.mapping.ambiguous ?? 0}`);
        lines.push(`- unmapped: ${data.mapping.unmapped ?? 0}`);
      }

      if (mappingBase.length) {
        lines.push('');
        lines.push('Mapping base');
        mappingBase.slice(0, 12).forEach((item, idx) => {
          lines.push(
            `${idx + 1}. ${item.source_column || 'Sin columna'} -> ${item.canonical_field || 'Sin campo canónico'} (${item.match_status || 'sin estado'})`
          );
        });
      }

      if (data.resumen_corto) {
        lines.push('');
        lines.push('Resumen corto');
        lines.push(`- ${data.resumen_corto}`);
      }

      if (previewCurado.length) {
        lines.push('');
        lines.push('Preview simple de salida curada');
        previewCurado.slice(0, 6).forEach((col, idx) => {
          lines.push(`${idx + 1}. ${col.source_column || 'Sin columna'} | ${col.canonical_field || 'Sin campo'} | tipo: ${col.value_type || 'n/a'}`);
          const samples = Array.isArray(col.samples) ? col.samples : [];
          samples.slice(0, 3).forEach((sample, sampleIdx) => {
            lines.push(
              `   - Muestra ${sampleIdx + 1}: "${sample.source_value ?? ''}" => "${sample.normalized_value ?? ''}" (confianza: ${sample.confidence ?? 'n/a'})`
            );
          });
        });
      }

      resultText.textContent = lines.join('\n');
      resultBox.style.display = 'block';
      renderObservationList(data);
    }

    function renderObservationList(data) {
      const details = Array.isArray(data.warning_details) ? data.warning_details : [];
      observationList.innerHTML = '';
      if (!details.length) {
        observationPanel.style.display = 'none';
        return;
      }
      observationPanel.style.display = 'block';
      details.forEach((obs, idx) => {
        const btn = document.createElement('button');
        btn.type = 'button';
        btn.className = 'obs-btn';
        const code = obs.code || `observación_${idx + 1}`;
        const severity = obs.severity ? ` · ${obs.severity}` : '';
        btn.textContent = `${code}${severity}`;
        btn.addEventListener('click', () => openObservationDetail(obs.observation_id));
        observationList.appendChild(btn);
      });
    }

    function openObservationDetail(observationId) {
      const details = Array.isArray(state.analysis?.warning_details) ? state.analysis.warning_details : [];
      const observation = details.find(item => item.observation_id === observationId);
      if (!observation) return;
      state.selectedObservationId = observationId;
      detailCode.textContent = observation.code || 'No disponible';
      detailCategory.textContent = observation.category || 'No disponible';
      detailSeverity.textContent = observation.severity || 'No disponible';
      detailMessage.textContent = observation.message || 'No disponible';
      detailField.textContent = observation.field_affected || 'No disponible';
      const examples = Array.isArray(observation.examples) ? observation.examples : [];
      detailExamples.textContent = examples.length ? examples.join(' | ') : 'No disponible';
      manualReviewNote.textContent = state.manualReview[observationId]
        ? 'Marcada para revisión manual (solo visual).'
        : '';
      showScreen3();
      pageTitle.textContent = 'Detalle de observación';
      pageSubtitle.textContent = 'Revise el warning seleccionado y decida si requiere intervención manual.';
      statusPill.textContent = 'Observación';
      statusPill.style.background = '#eef2f7';
      statusPill.style.color = '#30445d';
    }

    function renderErrorResult(message) {
      resultText.textContent = `Error en análisis: ${message || 'Error desconocido'}`;
      resultBox.style.display = 'block';
    }

    function renderRuleDetail(rule) {
      state.selectedRule = rule || null;
      ruleDetailCliente.textContent = rule?.cliente || 'No disponible';
      ruleDetailTipo.textContent = rule?.tipo || 'No disponible';
      ruleDetailCondicion.textContent = rule?.condicion || 'No disponible';
      ruleDetailAccion.textContent = rule?.accion || 'No disponible';
      const ejemplos = Array.isArray(rule?.ejemplos) ? rule.ejemplos : [];
      ruleDetailEjemplos.textContent = ejemplos.length ? ejemplos.join(' | ') : 'No disponible';
      ruleDetailImpacto.textContent = rule?.impacto_esperado || 'No disponible';
    }

    function renderRulesList(rows) {
      rulesList.innerHTML = '';
      if (!rows.length) {
        rulesList.innerHTML = '<div class="rules-empty">No hay reglas para este cliente.</div>';
        renderRuleDetail(null);
        return;
      }
      rows.forEach((rule) => {
        const row = document.createElement('div');
        row.className = 'rules-row';
        row.innerHTML = `
          <div><strong>${rule.nombre || 'Sin nombre'}</strong></div>
          <div>${rule.categoria || 'n/d'}</div>
          <div>${rule.origen || 'n/d'}</div>
          <div>${rule.estado || 'n/d'}</div>
          <div>${rule.ultimo_uso || 'n/d'}</div>
          <div><button type="button" class="secondary-btn">Abrir</button></div>
        `;
        row.querySelector('button')?.addEventListener('click', () => renderRuleDetail(rule));
        rulesList.appendChild(row);
      });
      renderRuleDetail(rows[0]);
    }

    async function fetchRules(clienteNombre='') {
      const params = new URLSearchParams();
      if (clienteNombre) params.set('cliente', clienteNombre);
      const query = params.toString();
      const response = await fetch(`/api/reglas${query ? `?${query}` : ''}`);
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || 'No se pudo cargar reglas.');
      return data;
    }

    async function postRuleAction(action, ruleId, payload={}) {
      const response = await fetch('/api/reglas/accion', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({action, rule_id: ruleId, payload}),
      });
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || 'No se pudo aplicar la acción de regla.');
      return data;
    }

    async function openRulesScreen() {
      showScreen4();
      rulesList.innerHTML = '<div class="rules-empty">Cargando reglas...</div>';
      const clienteNombre = cliente.value.trim() || 'Cliente general';
      try {
        const data = await fetchRules(clienteNombre);
        const rows = Array.isArray(data.reglas) ? data.reglas : [];
        state.rules = rows;
        renderRulesList(rows);
      } catch (err) {
        rulesList.innerHTML = `<div class="rules-empty">Error: ${err.message}</div>`;
        renderRuleDetail(null);
      }
    }

    async function reloadRulesKeepSelection() {
      const selectedId = state.selectedRule?.id;
      const clienteNombre = cliente.value.trim() || 'Cliente general';
      const data = await fetchRules(clienteNombre);
      const rows = Array.isArray(data.reglas) ? data.reglas : [];
      state.rules = rows;
      renderRulesList(rows);
      if (selectedId) {
        const selected = rows.find((r) => Number(r.id) === Number(selectedId));
        if (selected) renderRuleDetail(selected);
      }
    }

    function getSuggestedCloseStatus(data) {
      const warningCodes = Array.isArray(data?.warning_codes) ? data.warning_codes : [];
      const details = Array.isArray(data?.warning_details) ? data.warning_details : [];
      const hasCritical = details.some((w) => ['high', 'critical'].includes(String(w.severity || '').toLowerCase()));
      const hasManual = Object.keys(state.manualReview || {}).length > 0;
      if (hasCritical || hasManual) return { label: 'Requiere revisión manual', cls: 'status-critical' };
      if (warningCodes.length) return { label: 'Con observaciones', cls: 'status-warn' };
      return { label: 'Lista para exportación', cls: 'status-ok' };
    }

    function openCloseScreenWithData(data) {
      if (!data) return;
      state.closeContext = data;
      const warningCodes = Array.isArray(data.warning_codes) ? data.warning_codes : [];
      closeFile.textContent = data.file || data.archivo || 'No disponible';
      closeSheet.textContent = data.sheet || data.hoja_detectada || 'No disponible';
      closeHeader.textContent = data.header_row_1based || data.fila_header || 'No disponible';
      closeWarnings.textContent = warningCodes.length ? warningCodes.join(', ') : 'Sin warning_codes';
      closeSummary.textContent = data.resumen_corto || 'Sin resumen';
      const suggested = getSuggestedCloseStatus(data);
      closeStatusChip.textContent = suggested.label;
      closeStatusChip.className = `status-chip ${suggested.cls}`;
      showScreen6();
    }

    function triggerDownload(filename, content, contentType='text/plain;charset=utf-8') {
      const blob = new Blob([content], { type: contentType });
      const url = URL.createObjectURL(blob);
      const a = document.createElement('a');
      a.href = url;
      a.download = filename;
      document.body.appendChild(a);
      a.click();
      a.remove();
      URL.revokeObjectURL(url);
    }

    function renderHistoryList(rows) {
      historyList.innerHTML = '';
      if (!rows.length) {
        const empty = document.createElement('div');
        empty.className = 'history-empty';
        empty.textContent = 'No hay revisiones guardadas todavía.';
        historyList.appendChild(empty);
        return;
      }
      rows.forEach((row) => {
        const item = document.createElement('div');
        item.className = 'history-item';
        const warningCodes = Array.isArray(row.warning_codes) ? row.warning_codes.join(', ') : 'Sin warning_codes';
        item.innerHTML = `
          <div>
            <div class="history-title">${row.archivo || 'Sin archivo'}</div>
            <div class="history-meta">${row.fecha_hora || 'Sin fecha'} · Hoja: ${row.hoja_detectada || 'n/d'} · Header: ${row.fila_header || 'n/d'}</div>
            <div class="history-meta">warning_codes: ${warningCodes}</div>
            <div class="history-meta">Resumen: ${row.resumen_corto || 'Sin resumen'}</div>
          </div>
        `;
        const btn = document.createElement('button');
        btn.type = 'button';
        btn.className = 'secondary-btn';
        btn.textContent = 'Abrir revisión';
        btn.addEventListener('click', () => openSavedRevision(row));
        item.appendChild(btn);
        historyList.appendChild(item);
      });
    }

    function openSavedRevision(row) {
      state.selectedRevision = row;
      const warningCodes = Array.isArray(row.warning_codes) ? row.warning_codes : [];
      const lines = [
        `Archivo: ${row.archivo || 'No disponible'}`,
        `Fecha/hora: ${row.fecha_hora || 'No disponible'}`,
        `Hoja detectada: ${row.hoja_detectada || 'No disponible'}`,
        `Fila header: ${row.fila_header || 'No disponible'}`,
        `warning_codes: ${warningCodes.length ? warningCodes.join(', ') : 'Sin warning_codes'}`,
        `Resumen corto: ${row.resumen_corto || 'Sin resumen'}`,
      ];
      historyDetailText.textContent = lines.join('\n');
      historyDetail.style.display = 'block';
    }

    function openSavedRevisionInResult() {
      if (!state.selectedRevision) return;
      const row = state.selectedRevision;
      const analysisLike = {
        file: row.archivo || 'No disponible',
        sheet: row.hoja_detectada || null,
        header_row_1based: row.fila_header || null,
        warning_codes: Array.isArray(row.warning_codes) ? row.warning_codes : [],
        resumen_corto: row.resumen_corto || 'Sin resumen',
        mapping_base: [],
        preview_curado_simple: [],
        warning_details: Array.isArray(row.warning_details) ? row.warning_details : [],
      };
      showScreen2();
      renderCuratedResult(analysisLike);
      pageTitle.textContent = 'Resultado curado (guardado)';
      pageSubtitle.textContent = 'Vista de una revisión cargada desde el historial local.';
      statusPill.textContent = 'Revisión seleccionada';
      statusPill.style.background = '#dce9f7';
      statusPill.style.color = '#314960';
      setStatus('Revisión guardada abierta desde historial.', '');
    }

    function setEmptyState() {
      state.file = null;
      state.preview = null;
      state.valid = false;
      state.analysis = null;
      state.selectedObservationId = null;
      state.manualReview = {};
      state.rules = [];
      state.selectedRule = null;
      state.selectedRevision = null;
      state.closeContext = null;
      showScreen1();
      fileInput.value = '';
      fileCard.style.display = 'none';
      pageTitle.textContent = 'Sin archivo cargado';
      pageSubtitle.textContent = 'Seleccione un archivo Excel o CSV para iniciar la revisión.';
      statusPill.textContent = 'Esperando selección';
      statusPill.style.background = '#e8edf3';
      statusPill.style.color = '#4b5d72';
      analyzeBtn.disabled = true;
      setStatus('Listo para seleccionar un archivo.');
      renderChecks([
        {label:'Archivo detectado', state:'Pendiente', note:'Sin archivo'},
        {label:'Extensión válida', state:'Pendiente', note:'Excel o CSV'},
        {label:'Hojas detectadas', state:'Pendiente', note:'A la espera de archivo'},
        {label:'Listo para análisis', state:'Bloqueado', note:'Debe cargar un archivo'},
      ]);
      resultBox.style.display = 'none';
      resultText.textContent = '';
      observationPanel.style.display = 'none';
      observationList.innerHTML = '';
      manualReviewNote.textContent = '';
      historyDetail.style.display = 'none';
      historyDetailText.textContent = '';
      historyList.innerHTML = '';
      closeFinalNote.value = '';
      closeTraceNote.value = '';
    }

    async function sendPreview(file) {
      const form = new FormData();
      form.append('archivo', file);
      setStatus('Validando archivo...', '');
      const response = await fetch('/api/preview', {method:'POST', body: form});
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || 'No se pudo validar el archivo.');
      return data;
    }

    async function sendAnalyze(file) {
      const form = new FormData();
      form.append('archivo', file);
      form.append('cliente', cliente.value.trim());
      form.append('periodo_fiscal', periodoFiscal.value.trim());
      form.append('tipo_revision', tipoRevision.value);
      setStatus('Iniciando análisis...', '');
      const response = await fetch('/api/analyze', {method:'POST', body: form});
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || 'No se pudo iniciar el análisis.');
      return data;
    }

    async function fetchHistory() {
      const response = await fetch('/api/revisiones');
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || 'No se pudo cargar el historial.');
      return data;
    }

    async function openHistoryScreen() {
      showScreen5();
      historyDetail.style.display = 'none';
      historyDetailText.textContent = '';
      state.selectedRevision = null;
      historyList.innerHTML = '<div class="history-empty">Cargando historial...</div>';
      try {
        const data = await fetchHistory();
        const rows = Array.isArray(data.revisiones) ? data.revisiones : [];
        state.history = rows;
        renderHistoryList(rows);
      } catch (err) {
        historyList.innerHTML = `<div class="history-empty">Error: ${err.message}</div>`;
      }
    }

    async function handleFile(file) {
      showScreen1();
      state.file = file;
      pageTitle.textContent = file.name;
      pageSubtitle.textContent = 'Cliente y período fiscal listos para confirmar.';
      fileName.textContent = file.name;
      fileMeta.textContent = `${Math.max(1, Math.round(file.size / 1024))} KB · ${(file.name.split('.').pop() || '').toUpperCase()}`;
      fileCard.style.display = 'flex';
      try {
        const preview = await sendPreview(file);
        state.preview = preview;
        state.valid = preview.state === 'ready';
        statusPill.textContent = state.valid ? 'Pendiente de análisis' : 'Archivo inválido';
        statusPill.style.background = state.valid ? '#e8edf3' : '#f8e2e2';
        statusPill.style.color = state.valid ? '#4b5d72' : '#8f4c4c';
        renderChecks(preview.prechecks || []);
        analyzeBtn.disabled = !state.valid;
        setStatus(state.valid ? 'Archivo listo para completar metadatos.' : 'El archivo no cumple con la validación mínima.', state.valid ? '' : 'error');
      } catch (err) {
        state.valid = false;
        analyzeBtn.disabled = true;
        statusPill.textContent = 'Archivo inválido';
        statusPill.style.background = '#f8e2e2';
        statusPill.style.color = '#8f4c4c';
        renderChecks([
          {label:'Archivo detectado', state:'Error', note:file.name},
          {label:'Extensión válida', state:'Error', note:'No aceptado'},
          {label:'Hojas detectadas', state:'Error', note:'No disponible'},
          {label:'Listo para análisis', state:'Bloqueado', note:'Revise el archivo'},
        ]);
        setStatus(err.message, 'error');
      }
    }

    function resetSelection() {
      setEmptyState();
    }

    pickFileBtn.addEventListener('click', () => fileInput.click());
    newFileBtn.addEventListener('click', () => fileInput.click());
    removeFileBtn.addEventListener('click', resetSelection);
    clearBtn.addEventListener('click', resetSelection);
    backToCargaBtn.addEventListener('click', showScreen1);
    openCloseBtn.addEventListener('click', () => {
      const data = state.analysis || state.selectedRevision;
      if (!data) {
        setStatus('No hay revisión analizada para cerrar.', 'error');
        return;
      }
      openCloseScreenWithData(data);
    });
    backToResultBtn.addEventListener('click', () => {
      showScreen2();
      pageTitle.textContent = 'Resultado curado';
      pageSubtitle.textContent = 'Revise warnings, mapping base y preview curado del archivo analizado.';
      statusPill.textContent = 'Resultado curado';
      statusPill.style.background = '#dce9f7';
      statusPill.style.color = '#314960';
    });
    markManualBtn.addEventListener('click', () => {
      if (!state.selectedObservationId) return;
      state.manualReview[state.selectedObservationId] = true;
      manualReviewNote.textContent = 'Marcada para revisión manual (sin persistencia).';
    });
    historyNavBtn.addEventListener('click', openHistoryScreen);
    rulesNavBtn.addEventListener('click', openRulesScreen);
    rulesBackToCargaBtn.addEventListener('click', showScreen1);
    ruleToggleBtn.addEventListener('click', async () => {
      if (!state.selectedRule) return;
      try {
        await postRuleAction('toggle', Number(state.selectedRule.id));
        await reloadRulesKeepSelection();
        setStatus('Estado de regla actualizado.', '');
      } catch (err) {
        setStatus(err.message, 'error');
      }
    });
    ruleEditBtn.addEventListener('click', async () => {
      if (!state.selectedRule) return;
      const condicion = prompt('Editar condición de la regla:', state.selectedRule.condicion || '');
      if (condicion === null) return;
      const accion = prompt('Editar acción de la regla:', state.selectedRule.accion || '');
      if (accion === null) return;
      try {
        await postRuleAction('edit', Number(state.selectedRule.id), {condicion, accion});
        await reloadRulesKeepSelection();
        setStatus('Regla editada (modo simple).', '');
      } catch (err) {
        setStatus(err.message, 'error');
      }
    });
    ruleDuplicateBtn.addEventListener('click', async () => {
      if (!state.selectedRule) return;
      try {
        const response = await postRuleAction('duplicate', Number(state.selectedRule.id));
        await reloadRulesKeepSelection();
        if (response?.regla) renderRuleDetail(response.regla);
        setStatus('Regla duplicada.', '');
      } catch (err) {
        setStatus(err.message, 'error');
      }
    });
    ruleDeleteBtn.addEventListener('click', async () => {
      if (!state.selectedRule) return;
      const confirmDelete = confirm(`¿Eliminar la regla "${state.selectedRule.nombre || ''}"?`);
      if (!confirmDelete) return;
      try {
        await postRuleAction('delete', Number(state.selectedRule.id));
        await reloadRulesKeepSelection();
        setStatus('Regla eliminada.', '');
      } catch (err) {
        setStatus(err.message, 'error');
      }
    });
    historyBackToCargaBtn.addEventListener('click', showScreen1);
    historyBackToListBtn.addEventListener('click', () => {
      historyDetail.style.display = 'none';
      historyDetailText.textContent = '';
    });
    historyOpenInResultBtn.addEventListener('click', openSavedRevisionInResult);
    historyOpenCloseBtn.addEventListener('click', () => {
      if (!state.selectedRevision) return;
      openCloseScreenWithData(state.selectedRevision);
    });
    downloadCuratedBtn.addEventListener('click', () => {
      if (!state.closeContext) return;
      const fileBase = (state.closeContext.file || state.closeContext.archivo || 'revision_curada').replace(/\.[^.]+$/, '');
      const payload = {
        file: state.closeContext.file || state.closeContext.archivo || null,
        sheet: state.closeContext.sheet || state.closeContext.hoja_detectada || null,
        header_row_1based: state.closeContext.header_row_1based || state.closeContext.fila_header || null,
        warning_codes: Array.isArray(state.closeContext.warning_codes) ? state.closeContext.warning_codes : [],
        resumen_corto: state.closeContext.resumen_corto || null,
        warning_details: Array.isArray(state.closeContext.warning_details) ? state.closeContext.warning_details : [],
      };
      triggerDownload(`${fileBase}_salida_curada.json`, JSON.stringify(payload, null, 2), 'application/json;charset=utf-8');
      setStatus('Salida curada descargada en formato JSON local.', '');
    });
    downloadReviewSummaryBtn.addEventListener('click', () => {
      if (!state.closeContext) return;
      const warningCodes = Array.isArray(state.closeContext.warning_codes) ? state.closeContext.warning_codes : [];
      const summaryText = [
        `Archivo: ${state.closeContext.file || state.closeContext.archivo || 'No disponible'}`,
        `Hoja detectada: ${state.closeContext.sheet || state.closeContext.hoja_detectada || 'No disponible'}`,
        `Fila header: ${state.closeContext.header_row_1based || state.closeContext.fila_header || 'No disponible'}`,
        `warning_codes: ${warningCodes.length ? warningCodes.join(', ') : 'Sin warning_codes'}`,
        `Resumen corto: ${state.closeContext.resumen_corto || 'Sin resumen'}`,
        `Estado sugerido: ${closeStatusChip.textContent || 'No disponible'}`,
        `Observación final: ${closeFinalNote.value.trim() || 'Sin observación final'}`,
        `Comentario de trazabilidad: ${closeTraceNote.value.trim() || 'Sin comentario de trazabilidad'}`,
      ].join('\n');
      triggerDownload('resumen_revision.txt', summaryText, 'text/plain;charset=utf-8');
      setStatus('Resumen de revisión descargado en texto local.', '');
    });
    closeBackToHistoryBtn.addEventListener('click', openHistoryScreen);
    closeStartNewBtn.addEventListener('click', () => {
      setEmptyState();
      setStatus('Nueva revisión lista para iniciar.', '');
    });
    exampleBtn.addEventListener('click', () => dialog.showModal());
    closeDialog.addEventListener('click', () => dialog.close());
    downloadGuide.addEventListener('click', () => dialog.showModal());
    fileInput.addEventListener('change', (ev) => {
      const file = ev.target.files && ev.target.files[0];
      if (file) handleFile(file);
    });

    dropZone.addEventListener('dragover', (ev) => {
      ev.preventDefault();
      dropZone.classList.add('dragover');
    });
    dropZone.addEventListener('dragleave', () => dropZone.classList.remove('dragover'));
    dropZone.addEventListener('drop', (ev) => {
      ev.preventDefault();
      dropZone.classList.remove('dragover');
      const file = ev.dataTransfer.files && ev.dataTransfer.files[0];
      if (file) handleFile(file);
    });

    analyzeBtn.addEventListener('click', async () => {
      if (!state.file || !state.valid) return;
      analyzeBtn.disabled = true;
      try {
        const data = await sendAnalyze(state.file);
        showScreen2();
        renderCuratedResult(data);
        setStatus('Análisis listo para revisión humana.', '');
        pageTitle.textContent = 'Resultado curado';
        pageSubtitle.textContent = 'Revise warnings, mapping base y preview curado del archivo analizado.';
        statusPill.textContent = 'Resultado curado';
        statusPill.style.background = '#dce9f7';
        statusPill.style.color = '#314960';
      } catch (err) {
        setStatus(err.message, 'error');
        renderErrorResult(err.message);
      } finally {
        analyzeBtn.disabled = !state.valid;
      }
    });

    setEmptyState();
  </script>
</body>
</html>
""",
])
